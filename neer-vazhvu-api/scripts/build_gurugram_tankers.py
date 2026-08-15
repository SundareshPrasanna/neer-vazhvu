#!/usr/bin/env python3
"""
Build Gurugram's tanker series from GMDA's own bulk-water booking ledger.

WHY THIS IS A THIRD KIND OF TANKER DATA
---------------------------------------
The platform already holds two, and this is neither:

  * household-survey (Bengaluru) - OpenCity surveys of what households PAY a
    private, unregulated, RTI-gated market. Demand side, sampled, priced.
  * utility-ledger (Hyderabad) - HMWSSB's own record of bookings and
    DELIVERIES per division and section. Supply side, censused, and the
    config contract in src/lib/cities/types.ts says of it, correctly,
    "No prices exist in it."

Gurugram is a utility SALES ledger. GMDA dispenses bulk water by tanker from
six named points, charges for it by water type, and published the whole
transaction log for three years. So it has prices AND it is a census AND it
names the buyer - and it has no delivery confirmation at all, because a
booking here IS a collection. It answers a question neither existing page
asks: what did the utility SELL, to whom, and at what price?

Hence tankerDataKind 'utility-sales-ledger' and its own panel, per that same
contract's rule that the kinds "are not interchangeable and must not share a
renderer".

WHAT THE DATA ACTUALLY SHOWS (checked before designing the page)
---------------------------------------------------------------
The obvious metric - tanker volume over time - is a TRAP, and saying so is
part of the output. Bookings fall 12,337 -> 9,741 -> 7,208 across 2019-2021,
which reads as falling tanker dependence. It is not safe to read that way:
this is a construction-driven market and 2020-21 are COVID years. The volume
series ships with that caveat attached, not as a headline.

What survives the confound is COMPOSITION, because it is a ratio measured
inside the same disrupted period:

    non-potable share of litres    2019 29.7%  ->  2020 42.2%  ->  2021 51.2%

Potable volume falls 64% while treated and recycled volume holds roughly
flat. And the tariffs are stable to the decimal across all three years
(potable Rs 70.5/kL, recycle Rs 30.0/kL, CETP-treated Rs 8.0/kL), so this is
not a pricing artefact either. Over three years GMDA moved the majority of
its bulk tanker water off potable supply and onto treated effluent, in a city
the Central Ground Water Authority declared a dark zone in 2008.

LICENCE CONSTRAINT - WHY THIS AGGREGATES AND NEVER REPUBLISHES
--------------------------------------------------------------
gmda.gov.in asserts "All rights reserved" and publishes no reuse policy
(/terms-conditions.html and /disclaimer.html, both read 2026-08-14). Registry
entry `gmda-tanker-mis` records that posture. So this script emits counts,
sums and shares only. Two specific rules it enforces:

  * the delivery ADDRESS column is read (to count distinct sites) and then
    dropped. It is never written to the artifact.
  * buyers are emitted only as a ranked top-N of commercial entities by
    volume, which is the finding; the full 261-name list is not republished.

Source
------
GMDA static reports, one XLSX per year, at a path no index page links to:
    https://www.gmda.gov.in/static/report/Water%20Tanker%20MIS%20Report%20<YYYY>.xlsx

SOFT-404 TRAP: 2022 onward do NOT 404. They 302 to a 38,291-byte HTML error
page served as HTTP 200. Fetching by status code alone yields five phantom
editions of HTML parsed as a spreadsheet. `_fetch_year` asserts the ZIP magic
bytes before accepting a year, which is the same guard the Headwaters entry
applies via detection.expectContentType.

KNOWN GAPS, both upstream, both recorded in the output:
  - The series STOPS at 2021. Watched by `gmda-tanker-mis`; a 2022+ file
    appearing there is the event we want.
  - Two of the six dispensing stations appear in 2019 and then vanish
    (Water Works Chandu Bhudera, and the Raw/Canal Water type with it).
    Reported as a station-level gap rather than smoothed away.

Run
---
    cd neer-vazhvu-api
    python3 scripts/build_gurugram_tankers.py \
        --out ../public/data/gurugram-tanker-sales.json
"""

from __future__ import annotations

import argparse
import io
import re
import sys
import urllib.request
import zipfile
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

# The registry owns every registered source's licence string; a second copy in
# a generator is how the registry and the corpus drifted apart (PR #227).
sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "scripts"))
from registry_license import registry_license  # noqa: E402
from nvdm_write import write_artifact  # noqa: E402

SOURCE_ID = "gmda-tanker-mis"
URL = (
    "https://www.gmda.gov.in/static/report/Water%20Tanker%20MIS%20Report%20{year}.xlsx"
)
SOURCE_PAGE = "https://www.gmda.gov.in/onlineservices/water-tanker.html"
YEARS = (2019, 2020, 2021)

MONTHS = [
    "",
    "Jan",
    "Feb",
    "Mar",
    "Apr",
    "May",
    "Jun",
    "Jul",
    "Aug",
    "Sep",
    "Oct",
    "Nov",
    "Dec",
]

# Water types as GMDA labels them, in the order the page should read them:
# what the city drinks first, then the two treated alternatives it is
# substituting in. `Raw/Canal Water` exists only in 2019.
POTABLE = "Potable/Drinking Water"


def _fetch_year(year: int, timeout: int = 120) -> bytes | None:
    """Return the XLSX bytes for a year, or None if that year is not published.

    Status is not sufficient here. GMDA answers a missing edition with a 302
    to an HTML error page served as 200, so a status-only check accepts every
    year forever. XLSX is a ZIP container, so the two magic-byte prefixes for
    a non-empty zip are the cheap, format-level assertion that this is a
    spreadsheet and not a rendered error page.
    """
    req = urllib.request.Request(
        URL.format(year=year), headers={"User-Agent": "neervazhvu-ggm-tankers"}
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            body = resp.read()
    except Exception as exc:  # noqa: BLE001 - a dead year must not kill the run
        print(f"  {year}: fetch failed ({exc})", file=sys.stderr)
        return None
    if not body.startswith((b"PK\x03\x04", b"PK\x05\x06")):
        print(
            f"  {year}: not a spreadsheet ({len(body)} bytes) - soft-404",
            file=sys.stderr,
        )
        return None
    return body


def _norm_buyer(name: str) -> str:
    """Collapse the same commercial entity written several ways.

    The raw column holds 290 distinct strings for 261 actual buyers: "DLF City
    Centre Ltd" and "DLF City Centre Ltd sector 25A" are one buyer, as are
    "M/s JMC Project (India) Ltd" and "M/S JMC Projects (India) Limited". Drop
    the corporate-form noise and the trailing site descriptor, then key on the
    leading words. Deliberately conservative: over-merging two genuinely
    different developers would be worse than leaving both in the tail.
    """
    s = name.casefold()
    s = re.sub(
        r"\bm/?s\b|\bpvt\b|\bprivate\b|\bltd\b|\blimited\b|\bthe\b|\bllp\b", " ", s
    )
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    parts = [p for p in s.split() if p]
    return " ".join(parts[:3])


def _display_buyer(variants: dict[str, int]) -> str:
    """Pick the most-used raw spelling as the label for a merged buyer."""
    return max(variants.items(), key=lambda kv: (kv[1], -len(kv[0])))[0].strip()


def _rows(book: bytes, year: int) -> tuple[list[dict], int]:
    """Parse one year's workbook into normalised booking rows.

    Returns (rows, rejected). Rejects are counted and reported rather than
    dropped quietly - two rows in the three-year series are genuinely corrupt
    upstream and both would poison a total if coerced:

      * 2019-07-29, STP Behrampur: no buyer, 0 litres, and a tanker size of
        919911000000 - an epoch-millisecond value that has leaked out of the
        adjacent column.
      * 2020-03-13, Water Works Basai: an unescaped comma inside the delivery
        address has shifted the whole row left, so `buyer` reads "1", tanker
        size reads "Gurgaon" and amount reads 'Haryana"'.

    Column note: the sheet carries an unnamed second column holding the epoch
    milliseconds of the same timestamp already in `Date`. It is redundant and
    ignored. `Water Station` is the dispensing point and `Type` the water
    grade; both are labels GMDA controls, so they are carried verbatim rather
    than remapped.
    """
    import openpyxl  # imported here so --help works without the dependency

    wb = openpyxl.load_workbook(io.BytesIO(book), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(it)]
    idx = {name: i for i, name in enumerate(header)}

    required = [
        "Date",
        "Water Station",
        "Type",
        "Name",
        "Amount",
        "Total Water QTY Drawn in Liters",
    ]
    missing = [c for c in required if c not in idx]
    if missing:
        raise RuntimeError(
            f"{year}: workbook is missing columns {missing}; header was {header}"
        )

    out: list[dict] = []
    rejected = 0
    for raw in it:
        if raw is None or all(v is None for v in raw):
            continue
        get = lambda c: raw[idx[c]] if idx[c] < len(raw) else None  # noqa: E731

        when = get("Date")
        if isinstance(when, str):
            try:
                when = datetime.fromisoformat(when)
            except ValueError:
                rejected += 1
                continue
        if not isinstance(when, datetime):
            rejected += 1
            continue
        if when.year != year:
            # Defensive: a stray row from a neighbouring year would silently
            # double-count a month across two files.
            rejected += 1
            continue

        litres = get("Total Water QTY Drawn in Liters")
        amount = get("Amount")
        if not isinstance(litres, (int, float)) or litres <= 0:
            rejected += 1
            continue

        out.append(
            {
                "when": when,
                "station": str(get("Water Station") or "").strip(),
                "wtype": str(get("Type") or "").strip(),
                "buyer": str(get("Name") or "").strip(),
                # Read only to count distinct delivery sites. Never emitted -
                # see the licence note in the module docstring.
                "site": str(get("Address") or "").strip() if "Address" in idx else "",
                "litres": float(litres),
                "amount": float(amount) if isinstance(amount, (int, float)) else 0.0,
            }
        )
    return out, rejected


def _rate(rows: list[dict]) -> float | None:
    """Rupees per kilolitre over a set of bookings, or None if unpriced."""
    litres = sum(r["litres"] for r in rows)
    amount = sum(r["amount"] for r in rows)
    if litres <= 0 or amount <= 0:
        return None
    return round(amount / (litres / 1000.0), 2)


def build(rows: list[dict], years_present: list[int], rejected: dict[int, int]) -> dict:
    total_l = sum(r["litres"] for r in rows)
    total_rs = sum(r["amount"] for r in rows)

    by_year: list[dict] = []
    for y in years_present:
        yr = [r for r in rows if r["when"].year == y]
        nonpot = [r for r in yr if r["wtype"] != POTABLE]
        yl = sum(r["litres"] for r in yr)
        by_year.append(
            {
                "year": y,
                "bookings": len(yr),
                "litres": int(yl),
                "amount_inr": int(round(sum(r["amount"] for r in yr))),
                "buyers": len({_norm_buyer(r["buyer"]) for r in yr if r["buyer"]}),
                "non_potable_litres": int(sum(r["litres"] for r in nonpot)),
                "non_potable_pct": round(
                    100.0 * sum(r["litres"] for r in nonpot) / yl, 1
                )
                if yl
                else None,
                # Malformed upstream rows refused on parse. Reported per year
                # so a future edition that arrives badly broken is visible as
                # a number rather than as a quietly smaller total.
                "rows_rejected": rejected.get(y, 0),
            }
        )

    # Water type x year: the composition finding, and the stable tariff that
    # rules out a pricing explanation for it.
    types: list[dict] = []
    for wtype in sorted({r["wtype"] for r in rows if r["wtype"]}):
        sel = [r for r in rows if r["wtype"] == wtype]
        types.append(
            {
                "water_type": wtype,
                "potable": wtype == POTABLE,
                "bookings": len(sel),
                "litres": int(sum(r["litres"] for r in sel)),
                "amount_inr": int(round(sum(r["amount"] for r in sel))),
                "rate_inr_per_kl": _rate(sel),
                "by_year": [
                    {
                        "year": y,
                        "litres": int(
                            sum(r["litres"] for r in sel if r["when"].year == y)
                        ),
                        "rate_inr_per_kl": _rate(
                            [r for r in sel if r["when"].year == y]
                        ),
                    }
                    for y in years_present
                ],
            }
        )
    types.sort(key=lambda t: -t["litres"])

    # Monthly series. Carried because seasonality is real (this is northern
    # India, and construction water tracks the building season), but the page
    # must read it with the COVID caveat that `_caveats` carries.
    monthly: list[dict] = []
    buckets: dict[tuple[int, int], list[dict]] = defaultdict(list)
    for r in rows:
        buckets[(r["when"].year, r["when"].month)].append(r)
    for y, m in sorted(buckets):
        sel = buckets[(y, m)]
        nonpot = sum(r["litres"] for r in sel if r["wtype"] != POTABLE)
        ml = sum(r["litres"] for r in sel)
        monthly.append(
            {
                "month": f"{y}-{m:02d}",
                "label": f"{MONTHS[m]} {y}",
                "bookings": len(sel),
                "litres": int(ml),
                "non_potable_pct": round(100.0 * nonpot / ml, 1) if ml else None,
            }
        )

    # Mean litres per calendar month across the series - the seasonal shape,
    # with the year-count exposed so a partial month is visible as such.
    seasonality: list[dict] = []
    for m in range(1, 13):
        sel = [b for b in buckets if b[1] == m]
        if not sel:
            continue
        seasonality.append(
            {
                "month": m,
                "label": MONTHS[m],
                "mean_litres": int(
                    round(
                        sum(sum(r["litres"] for r in buckets[k]) for k in sel)
                        / len(sel)
                    )
                ),
                "years": len(sel),
            }
        )

    stations: list[dict] = []
    for st in sorted({r["station"] for r in rows if r["station"]}):
        sel = [r for r in rows if r["station"] == st]
        yrs = sorted({r["when"].year for r in sel})
        stations.append(
            {
                "station": st,
                "bookings": len(sel),
                "litres": int(sum(r["litres"] for r in sel)),
                "water_types": sorted({r["wtype"] for r in sel if r["wtype"]}),
                "years": yrs,
                # A station that stops reporting is a fact about the network,
                # not a rounding error. Surfaced rather than smoothed.
                "active_all_years": yrs == years_present,
            }
        )
    stations.sort(key=lambda s: -s["litres"])

    # Top buyers only. The full name list is not republished (licence note in
    # the module docstring); the finding is the concentration and the sector,
    # both of which a ranked head carries.
    merged: dict[str, dict] = {}
    for r in rows:
        if not r["buyer"]:
            continue
        k = _norm_buyer(r["buyer"])
        e = merged.setdefault(
            k,
            {
                "litres": 0.0,
                "bookings": 0,
                "variants": defaultdict(int),
                "potable": 0.0,
            },
        )
        e["litres"] += r["litres"]
        e["bookings"] += 1
        e["variants"][r["buyer"]] += 1
        if r["wtype"] == POTABLE:
            e["potable"] += r["litres"]
    top_buyers = [
        {
            "buyer": _display_buyer(e["variants"]),
            "bookings": e["bookings"],
            "litres": int(e["litres"]),
            "share_pct": round(100.0 * e["litres"] / total_l, 2) if total_l else None,
            "potable_pct": round(100.0 * e["potable"] / e["litres"], 1)
            if e["litres"]
            else None,
        }
        for e in sorted(merged.values(), key=lambda x: -x["litres"])[:15]
    ]

    first, last = min(r["when"] for r in rows), max(r["when"] for r in rows)
    # NVDM v1 envelope, emitted BY THE PRODUCER rather than by a separate
    # per-city injector. That is the pattern the Madurai pilot settled on and
    # the one the Hyderabad injector's own docstring flags as preferred: a
    # regenerating producer owns its envelope, so a re-run cannot strip it.
    # Cities 1-6 needed injectors only because their artifacts predated NVDM.
    envelope = {
        "nvdm": "1.0",
        # City lives in `scope`, so the dataset id is the stem WITHOUT it -
        # the same shape as data-root/rainfall-recent. The catalogue derives
        # this from the path and validate_nvdm asserts the two agree.
        "dataset": "data-root/tanker-sales",
        "scope": {"kind": "city", "id": "gurugram"},
        "provenance": {
            "sources": [
                {
                    "id": SOURCE_ID,
                    "title": "GMDA bulk-water tanker booking ledger (Water Tanker MIS Report, one XLSX per year)",
                    "publisher": "Gurugram Metropolitan Development Authority",
                    "license": registry_license(SOURCE_ID),
                }
            ],
            # "scrape": the workbooks are fetched from a static path on the
            # publisher's site. Not "api" - there is no service here, just a
            # file whose URL we worked out.
            "method": "scrape",
            "produced_at": date.today().isoformat(),
            "produced_by": "neer-vazhvu-api/scripts/build_gurugram_tankers.py",
            "note": (
                "Aggregated on build, never republished row-for-row: the publisher asserts all "
                "rights reserved and states no reuse policy, so the artifact carries counts, sums "
                "and shares only, the delivery-address column is dropped, and buyers appear solely "
                "as a ranked top-15 by volume. Two corrupt upstream rows were refused rather than "
                "coerced; see _caveats and by_year[].rows_rejected."
            ),
        },
    }
    return {
        **envelope,
        "_source": "Gurugram Metropolitan Development Authority - Water Tanker MIS Report",
        "_source_url": SOURCE_PAGE,
        "_licence": registry_license(SOURCE_ID),
        "_fetched": date.today().isoformat(),
        "_note": (
            "GMDA's own bulk-water tanker booking ledger, published as one XLSX per year. "
            "Every booking is a collection from a GMDA dispensing point, so unlike Hyderabad's "
            "utility ledger there is no separate delivery confirmation to measure, and unlike "
            "Bengaluru's household survey the prices here are the utility's own tariff rather "
            "than a private market's. Aggregated on build: GMDA asserts all rights reserved and "
            "publishes no reuse policy, so no upstream row is republished and the delivery-address "
            "column is dropped entirely."
        ),
        "_coverage_gap": (
            "The published series stops after 2021. Verified 2026-08-14: 2022 through 2026 do not "
            "404 but redirect to an HTML error page served as HTTP 200, so they are absent rather "
            "than empty. Watched by the Headwaters entry gmda-tanker-mis."
        ),
        "_caveats": [
            "Do not read the falling volume as falling tanker dependence. Bookings drop 12,337 to "
            "7,208 across 2019-2021, but this is a construction-driven market and 2020-21 are "
            "COVID years. The composition trend below is the finding that survives, because it is "
            "a ratio measured inside the same disrupted period.",
            "Buyer counts are after normalising corporate-form and site-suffix variants of the "
            "same name; the raw column carries more distinct strings than there are buyers.",
            "Two upstream rows across the three years are corrupt and were refused rather than "
            "coerced: one carries an epoch-millisecond value in the tanker-size column with zero "
            "litres, the other has an unescaped comma in the delivery address that shifts the "
            "whole row. Counted per year as rows_rejected.",
        ],
        "totals": {
            "bookings": len(rows),
            "litres": int(total_l),
            "amount_inr": int(round(total_rs)),
            "years": len(years_present),
            "months": len(monthly),
            "buyers": len(merged),
            "delivery_sites": len({r["site"] for r in rows if r["site"]}),
            "stations": len(stations),
            "first_booking": first.date().isoformat(),
            "last_booking": last.date().isoformat(),
            "non_potable_pct_first_year": by_year[0]["non_potable_pct"]
            if by_year
            else None,
            "non_potable_pct_last_year": by_year[-1]["non_potable_pct"]
            if by_year
            else None,
        },
        "by_year": by_year,
        "water_types": types,
        "monthly": monthly,
        "seasonality": seasonality,
        "stations": stations,
        "top_buyers": top_buyers,
    }


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--out", required=True, help="write JSON here")
    args = ap.parse_args()

    rows: list[dict] = []
    present: list[int] = []
    rejected: dict[int, int] = {}
    for y in YEARS:
        print(f"fetching {y}...", file=sys.stderr)
        book = _fetch_year(y)
        if book is None:
            continue
        try:
            got, bad = _rows(book, y)
        except (zipfile.BadZipFile, RuntimeError) as exc:
            print(f"  {y}: parse failed ({exc})", file=sys.stderr)
            continue
        if not got:
            print(f"  {y}: zero usable rows", file=sys.stderr)
            continue
        print(f"  {y}: {len(got)} bookings ({bad} rejected)", file=sys.stderr)
        rows.extend(got)
        rejected[y] = bad
        present.append(y)

    if not rows:
        print("no usable rows from any year - refusing to write", file=sys.stderr)
        return 1

    out = build(rows, sorted(present), rejected)
    write_artifact(Path(args.out), out)
    t = out["totals"]
    print(
        f"wrote {args.out}: {t['bookings']} bookings, {t['litres']:,} L, "
        f"Rs {t['amount_inr']:,}, {t['months']} months, {t['buyers']} buyers",
        file=sys.stderr,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
